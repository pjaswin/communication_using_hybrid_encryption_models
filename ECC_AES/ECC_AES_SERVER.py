import socket
import pickle
import time
import numpy as np
import openpyxl
from tinyec import registry
from encrypts import AES
import hashlib, secrets, binascii
import tracemalloc
import psutil,os


def hamming_distance(s1, s2):
    return sum(ch1 != ch2 for ch1, ch2 in zip(s1, s2))
# def ecc_point_to_256_bit_key(point):
#     sha = hashlib.sha256(int.to_bytes(point.x, 32, 'big'))
#     sha.update(int.to_bytes(point.y, 32, 'big'))
#     return sha.digest()
def ecc_point_to_128_bit_key(point):
    sha = hashlib.sha256(int.to_bytes(point.x, 32, 'big'))
    sha.update(int.to_bytes(point.y, 32, 'big'))
    key_256_bit = sha.digest()
    key_128_bit = key_256_bit[:16]  # Extract the first 16 bytes (128 bits) of the 256-bit key
    return key_128_bit

# Create a socket object
serversocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

# Get local machine name
host = socket.gethostname()

port = 9999

# Bind to a specific IP and port
serversocket.bind((host, port))

# Listen for incoming connections
serversocket.listen(5)

clientsocket,addr = serversocket.accept()
print("Got a connection from %s" % str(addr))
workbook1 = openpyxl.Workbook()
sheet1 = workbook1.active
# workbook2 = openpyxl.Workbook()
# sheet2 = workbook2.active
file_count=0
file_list=["onekb.txt","twokb.txt","fivekb.txt","tenkb.txt","twentykb.txt"]
dup_list=["testonekb.txt","testtwokb.txt","testfivekb.txt","testtenkb.txt","testtwentykb.txt"]
outputs=["file name","encrption time","decryption time","Cpu encryption","CPU decryption","memory encryption","memory decryption", "throughput","Hamming Dist","Avalanche Effect","Correlation Coff"]
for column, output in enumerate(outputs, start=1):
    sheet1.cell(row=1, column=column, value=output)
# outputs=["file name","encrption time","decryption time","Cpu encryption","CPU decryption","memory encryption","memory decryption"]
# for column, output in enumerate(outputs, start=1):
#     sheet2.cell(row=1, column=column, value=output)
row_value=2
# obj=Blowfish2.Blowfish()
# rsa1=RSA.RSA()
while file_count<5:
    round_count = 0
    while round_count<30:
        process = psutil.Process()
        start_cpu_key= process.cpu_percent()
        tracemalloc.start()
        current_key1, peak_key1 = tracemalloc.get_traced_memory()
        start_time= time.perf_counter()
        msg = clientsocket.recv(1024)
        clientpubkey = pickle.loads(msg)
        curve = registry.get_curve('secp256r1')
        privKey = secrets.randbelow(curve.field.n)
        pubKey = privKey * curve.g #blowfish key generation
        data=pickle.dumps(pubKey)
        clientsocket.send(data)
        sharedkey=privKey*clientpubkey
        secretKey = ecc_point_to_128_bit_key(sharedkey)
        aes1=AES.AES(secretKey)
        end_time= time.perf_counter()
        current_key2, peak_key2 = tracemalloc.get_traced_memory()
        time_taken1=end_time-start_time
        # end_cpu_key= psutil.cpu_percent()
        # time_taken=end_time_key-start_time_key
        # cpu_usage=end_cpu_key-start_cpu_key
        # mem_usage=current_key2-current_key1
        # print("Time take for key exchange and gen",time_taken)
        # print("CPU for key exchange and gen",cpu_usage)
        # print("Mem for key exchange and gen",mem_usage)
        # msg = clientsocket.recv(1024 * 100)
        # samplelist = pickle.loads(msg)
        # outputs = [file_list[file_count], time_taken,samplelist[1], cpu_usage,samplelist[0], mem_usage,samplelist[2]]
        # for column, output in enumerate(outputs, start=1):
        #     sheet2.cell(row=row_value, column=column, value=output)

        with open(file_list[file_count], 'r') as file:
            sub = file.read()
        start_time = time.perf_counter()
        modified_msg = sub[:2] + 'XY' + sub[4:]
        hashed_word = hashlib.sha256(sub.encode()).hexdigest()
        x=aes1.Encrypt(sub)
        y=aes1.Encrypt(hashed_word)
        data=pickle.dumps(x)
        clientsocket.send(data)
        data1=pickle.dumps(y)
        clientsocket.send(data1)
        end_time= time.perf_counter()
        current_key, peak_key = tracemalloc.get_traced_memory()
        end_cpu= process.cpu_percent()
        time_taken2=end_time-start_time
        cpu_usage=end_cpu-start_cpu_key
        mem_usage=current_key-current_key1
        size=len(sub)*8
        with open(dup_list[file_count], 'r') as file:
            sub2 = file.read()
        x2=aes1.Encrypt(sub2)
        binary_string1 = bin(int.from_bytes(x, byteorder='big'))
        binary_string2 = bin(int.from_bytes(x2, byteorder='big'))
        q = hamming_distance(binary_string1, binary_string2)
        avalanche_eff=q / size
        # print(f"avalanche effect is {q / size}")
        # print(f"throughput is {size/time_taken_text}")
        #
        plaintext_data = np.array([ord(c) for c in sub])
        ciphertext_data = np.array([ord(c) for c in x.hex()])
        output_string1 = np.array2string(plaintext_data, separator=', ')
        output_list1 = []
        for x in output_string1[1:-1].split(','):
            try:
                output_list1.append(int(x))
            except ValueError:
                print(x)
        # output_list1 = [int(x) for x in output_string1[1:-1].split(',')]
        output_string2 = np.array2string(ciphertext_data, separator=', ')
        output_list2 = []
        for x in output_string2[1:-1].split(','):
            try:
                output_list2.append(int(x))
            except ValueError:
                print(x)
        # output_list2 = [int(x) for x in output_string2[1:-1].split(',')]
        # print(output_list1,output_list2)
        l=len(output_list1)
        co=np.corrcoef(output_list1, output_list2[:l])[0][1]
        msg = clientsocket.recv(1024 * 100)
        samplelist = pickle.loads(msg)
        # print(samplelist)
        time_taken=time_taken1+time_taken2
        print("time taken",time_taken)
        outputs = [file_list[file_count], time_taken,samplelist[1], cpu_usage,samplelist[0], mem_usage,samplelist[2],size/time_taken2,q,avalanche_eff,co]
        for column, output in enumerate(outputs, start=1):
            sheet1.cell(row=row_value, column=column, value=output)
        row_value=row_value+1
        round_count = round_count + 1
        time.sleep(5)
        print(f"file is {file_list[file_count]} and round is{round_count}")
    row_value=row_value+10
    file_count=file_count+1
workbook1.save('ECC_AES.xlsx')
clientsocket.close()





